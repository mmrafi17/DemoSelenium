import openpyxl

book1 = openpyxl.load_workbook('C:\\Users\\MRF\\Downloads\\download.xlsx')
sheet = book1.active
cell1 = sheet.cell(row=1, column=2)
dictionaryData = {}

dictionaryData1 = {}

# print(cell1.value)
# print(sheet.max_row)
# print(sheet.max_column)

for i in range(1, sheet.max_column+1):
    if sheet.cell(row=1, column=i).value == 'season':
        dictionaryData['col'] = i

for j in range(1, sheet.max_row+1):
    for k in range(1, sheet.max_column+1):
        if sheet.cell(row=j, column=k).value == 'Apple':
            dictionaryData['row'] = j


print(dictionaryData['col'])
print(dictionaryData['row'])
