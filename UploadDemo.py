import time

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
import openpyxl


def update_excel_file(filePath, searchTerm,  colName, newValue):
    book1 = openpyxl.load_workbook(filePath)
    sheet = book1.active
    dictionaryData = {}

    for i in range(1, sheet.max_column+1):
        if sheet.cell(row=1, column=i).value == colName:
            dictionaryData['col'] = i

    for j in range(1, sheet.max_row+1):
        for k in range(1, sheet.max_column+1):
            if sheet.cell(row=j, column=k).value == searchTerm:
                dictionaryData['row'] = j

    sheet.cell(row=dictionaryData['row'], column=dictionaryData['col']).value = newValue
    book1.save(file_path)


file_path = "C:/Users/MRF/Downloads/download.xlsx"
driver = webdriver.Chrome()
driver.implicitly_wait(5)
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.maximize_window()
driver.find_element(By.CSS_SELECTOR, "#downloadButton").click()

file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
file_input.send_keys(file_path)
toast_locator = (By.CSS_SELECTOR, ".Toastify__toast-body div:nth-child(2)")

wait = WebDriverWait(driver, 5)
wait.until(expected_conditions.visibility_of_element_located(toast_locator))
print(driver.find_element(*toast_locator).text)


fruitName = 'Apple'
priceUpdate = 'price'
valueUpdate = '889'
update_excel_file(file_path, fruitName, priceUpdate, valueUpdate)

priceColumn = driver.find_element(By.XPATH, "//div[text()='Price']").get_attribute("data-column-id")
actualPrice = driver.find_element(By.XPATH, "//div[text()='"+fruitName+"']/parent::div/parent::div/div[@id='cell-"+priceColumn+"-undefined']").text

assert valueUpdate == actualPrice

time.sleep(2)

